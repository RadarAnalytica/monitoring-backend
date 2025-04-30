import asyncio
from datetime import date
from io import BytesIO
from json import JSONDecodeError

from aiohttp import ClientSession, ContentTypeError, client_exceptions
from clickhouse_db.get_async_connection import get_async_connection
from server.utils.month_names import MONTH_NAMES
from server.utils.xl_header import make_radar_header
from settings import SEARCH_URL

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


async def get_query_small_data(
    http_session: ClientSession, query_string, dest=-1257786, limit=3, page=1, rqa=5, timeout=5
):
    _data = {"data": {"products": []}}
    counter = 0
    while len(_data.get("data", dict()).get("products", [])) < 2 and counter < rqa:
        counter += 1
        try:
            async with http_session.get(
                url=SEARCH_URL,
                params={
                    "resultset": "catalog",
                    "query": query_string,
                    "limit": limit,
                    "dest": dest,
                    "page": page,
                },
                timeout=timeout,
            ) as response:
                if response.ok:
                    try:
                        _data = await response.json(content_type="text/plain")
                    except (ContentTypeError, JSONDecodeError):
                        return _data
                else:
                    # logger.critical("response not ok")
                    continue
        except (TypeError, asyncio.TimeoutError) as e:
            # logger.critical(f"ОШИБКА, {type(e)}")
            continue
        except client_exceptions.ServerDisconnectedError:
            counter -= 1
            continue

    return _data


def unnest_subjects_list(subjects_list: list):
    result = dict()
    for subject_data in subjects_list:
        s_id = subject_data.get("id")
        s_name = subject_data.get("name", "")
        children = subject_data.get("childs", [])
        result[s_id] = s_name
        result.update(unnest_subjects_list(children))
    return result


async def get_today_subjects_dict(http_session):
    url = "https://static-basket-01.wbcontent.net/vol0/data/subject-base.json"
    async with http_session.get(url) as resp:
        result = await resp.json()
    subjects_dict = unnest_subjects_list(result)
    return subjects_dict


async def get_valid_products(products_list: list[int]):
    result = dict()
    try:
        async with ClientSession() as session:
            async with session.post(
                url="https://radarmarket.ru/api/monitoring-reports/products-external",
                json={
                    "products": products_list
                }
            ) as resp:
                result = await resp.json()
    except:
        pass
    return result


async def get_report_data(
    http_session: ClientSession, query_string
):
    data = await get_query_small_data(http_session=http_session, query_string=query_string)
    products: list = data.get("data", dict()).get("products", list())
    total = data.get("data", dict()).get("total")
    first_product = products[0]
    subject = first_product.get("subjectId", 0)
    return total, subject

async def get_report_dataset():
    stmt = r"""SELECT r.query, 
        rfn.fs,
        rfn.diff,
        round(rfn.growth, 2)
    FROM request as r 
    JOIN (
        SELECT 
            rf1.query_id, 
            rf1.fs, 
            ((rf1.fs * 100 / if(coalesce(rf2.fs, 0) = 0, 1, rf2.fs)) - 100) as growth,
            rf1.fs - coalesce(rf2.fs, 0) as diff
        FROM (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM 
                request_frequency 
            WHERE 
                date BETWEEN today() - 30 and yesterday() 
            GROUP BY query_id
        ) as rf1 
        JOIN (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM request_frequency 
            WHERE 
                date BETWEEN today() - 60 AND today() - 31 
            GROUP BY query_id
        ) as rf2 
        ON rf1.query_id = rf2.query_id 
        ORDER BY growth DESC
    ) as rfn 
    ON rfn.query_id = r.id 
    WHERE 
        rfn.fs > 10000 
    AND 
        match(r.query, '^[\s]*[0-9]{5,10}[\s]*$') 
    AND 
        rfn.growth >= 100
    ORDER BY 
        rfn.fs"""
    async with get_async_connection() as client:
        q = await client.query(stmt)
        result = list(q.result_rows)
    p_ids = [int(row[0]) for row in result]
    valid_products = await get_valid_products(products_list=p_ids)
    dataset = []
    for row in result:
        p_id = row[0].strip()
        valid_product = valid_products.get(p_id)
        if not valid_product:
            continue
        name = valid_product.get("name")
        brand = valid_product.get("brand")
        supplier = valid_product.get("supplier")
        subject = valid_product.get("subject")
        frequency = row[1] or 0
        frequency_diff = row[2] or 0
        frequency_growth = row[3] or 0
        revenue = valid_product.get("revenue")
        revenue_diff = valid_product.get("revenue_diff")
        revenue_growth = valid_product.get("revenue_growth")
        orders = valid_product.get("orders")
        orders_diff = valid_product.get("orders_diff")
        orders_growth = valid_product.get("orders_growth")
        price = valid_product.get("price")
        price_diff = valid_product.get("price_diff")
        price_growth = valid_product.get("price_growth")
        dataset.append((
            p_id,
            name,
            brand,
            supplier,
            subject,
            frequency,
            frequency_diff,
            round(frequency_growth, 2),
            round(revenue, 2),
            round(revenue_diff, 2),
            round(revenue_growth, 2),
            orders,
            orders_diff,
            round(orders_growth, 2),
            round(price, 2),
            round(price_diff, 2),
            round(price_growth, 2),
        ))
    return dataset


def create_file_from_dataset(dataset: list[tuple]):
    today = date.today()
    wb = Workbook()
    ws = wb.active
    make_radar_header(
        ws=ws,
        sheet_title="Топ 50",
        name=f"Топ 50 товаров, которые выросли с внешним трафиком в {MONTH_NAMES[today.month]} {today.year}г.",
    )

    ws["A4"].value = "Артикул"
    ws["B4"].value = "Наименование товара"
    ws["C4"].value = "Бренд"
    ws["D4"].value = "Поставщик"
    ws["E4"].value = "Предмет"
    ws["F4"].value = "Частота запросов"
    ws["G4"].value = "Изменение частоты запросов"
    ws["H4"].value = "% изменения частоты запросов"
    ws["I4"].value = "Выручка"
    ws["J4"].value = "Изменение выручки"
    ws["K4"].value = "% изменения выручки"
    ws["L4"].value = "Заказы"
    ws["M4"].value = "Изменение кол-ва заказов"
    ws["N4"].value = "% изменения кол-ва заказов"
    ws["O4"].value = "Цена"
    ws["P4"].value = "Изменение цены"
    ws["Q4"].value = "% изменения цены"

    header_fill = PatternFill(start_color="a653ec", end_color="a653ec", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Применяем стили к заголовкам
    for col in range(1, 18):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    column_widths = {
        "A": 15, # "Артикул"
        "B": 40, # "Наименование товара"
        "C": 20, # "Бренд"
        "D": 25, # "Поставщик"
        "E": 20, # "Предмет"
        "F": 20, # "Частота запросов"
        "G": 20, # "Изменение частоты запросов"
        "H": 20, # "% изменения частоты запросов"
        "I": 20, # "Выручка"
        "J": 20, # "Изменение выручки"
        "K": 20, # "% изменения выручки"
        "L": 20, # "Заказы"
        "M": 20, # "Изменение кол-ва заказов"
        "N": 20, # "% изменения кол-ва заказов"
        "O": 20, # "Цена"
        "P": 20, # "Изменение цены"
        "Q": 20, # "% изменения цены"
    }

    ws.row_dimensions[4].height = 40
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = f"A4:Q4"
    for row in dataset:
        ws.append(row)

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    wb.close()
    return excel_stream


async def get_external_report_download_bytes():
    dataset = await get_report_dataset()
    file = create_file_from_dataset(dataset=dataset)
    return file


def test():
    test_data = [
        tuple(i for i in range(17))
    ]
    f = create_file_from_dataset(test_data)
    with open("test.xlsx", "wb") as file:
        file.write(f.read())


if __name__ == "__main__":
    test()
