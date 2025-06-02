from openpyxl.worksheet.worksheet import Worksheet
from datetime import date

from settings import BASE_DIR

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image


def make_radar_header(
    ws: Worksheet,
    sheet_title="Sheet1",
    name="Отчёт",
    ad="Отчет сгенерирован с помощью сервиса Radar-Analytica. "
    "После регистрации будет доступен тестовый период - 3 дня.",
):
    ws.title = sheet_title
    today = date.today()
    month_start = today.replace(day=1, month=5)
    img = Image(BASE_DIR / "static" / "icon.png")

    img.height = 120
    img.width = 120

    ws.add_image(img, "A1")

    ws.merge_cells("A2:H2")
    ws["A2"] = name
    ws["A2"].font = Font(bold=True, color="FFFFFF", sz=15)  # Белый текст
    ws["A2"].fill = PatternFill(
        start_color="a653ec", end_color="a653ec", fill_type="solid"
    )  # Синий фон
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:H1")
    font_cambria_italic_underline = Font(
        bold=True, color="000000", sz=15, underline="single"
    )  # Белый текст
    ws["A1"].hyperlink = "https://radar-analytica.ru/"
    ws["A1"].value = ad
    ws["A1"].style = "Hyperlink"
    ws["A1"].fill = PatternFill(
        start_color="abcdef", end_color="abcdef", fill_type="solid"
    )  # Синий фон
    ws["A1"].font = font_cambria_italic_underline
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:D3")
    ws["A3"] = (
        f"Дата начала отчёта: {month_start.strftime('%d.%m.%Y')} / Дата формирования отчёта: {today.strftime('%d.%m.%Y')}"
    )
    ws["A3"].font = Font(color="000000")

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 50
