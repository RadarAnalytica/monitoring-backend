import asyncio
from datetime import date, timedelta
from io import BytesIO
from json import JSONDecodeError

from aiohttp import ClientSession, ContentTypeError, client_exceptions
from clickhouse_db.get_async_connection import get_async_connection
from server.utils.month_names import MONTH_NAMES
from server.utils.xl_header import make_radar_header
from settings import SEARCH_URL

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


async def get_query_small_data(
    http_session: ClientSession,
    query_string,
    dest=-1257786,
    limit=3,
    page=1,
    rqa=5,
    timeout=5,
):
    _data = {"data": {"products": []}}
    counter = 0
    while len(_data.get("data", dict()).get("products", [])) < 2 and counter < rqa:
        counter += 1
        try:
            async with http_session.get(
                url=SEARCH_URL,
                params={
                    "resultset": "catalog",
                    "query": query_string,
                    "limit": limit,
                    "dest": dest,
                    "page": page,
                },
                timeout=timeout,
            ) as response:
                if response.ok:
                    try:
                        _data = await response.json(content_type="text/plain")
                    except (ContentTypeError, JSONDecodeError):
                        return _data
                else:
                    # logger.critical("response not ok")
                    continue
        except (TypeError, asyncio.TimeoutError) as e:
            # logger.critical(f"ОШИБКА, {type(e)}")
            continue
        except client_exceptions.ServerDisconnectedError:
            counter -= 1
            continue

    return _data


def unnest_subjects_list(subjects_list: list):
    result = dict()
    for subject_data in subjects_list:
        s_id = subject_data.get("id")
        s_name = subject_data.get("name", "")
        children = subject_data.get("childs", [])
        result[s_id] = s_name
        result.update(unnest_subjects_list(children))
    return result


async def get_today_subjects_dict(http_session):
    url = "https://static-basket-01.wbcontent.net/vol0/data/subject-base.json"
    async with http_session.get(url) as resp:
        result = await resp.json()
    subjects_dict = unnest_subjects_list(result)
    return subjects_dict


async def get_report_data(http_session: ClientSession, query_string):
    data = await get_query_small_data(
        http_session=http_session, query_string=query_string
    )
    products: list = data.get("data", dict()).get("products", list())
    total = data.get("data", dict()).get("total")
    first_product = products[0]
    subject = first_product.get("subjectId", 0)
    return total, subject


async def get_report_dataset(date_: str|date):
    stmt = r"""SELECT r.query, 
        rfn.fs, 
        round(rfn.growth, 2), 
        round(rfn60.growth, 2), 
        round(rfn90.growth, 2) 
    FROM (select * from request final where quantity > 10000 and NOT match(query, '^[\s]*[0-9]+[\s]*$')) as r 
    JOIN (
        SELECT 
            rf1.query_id, 
            rf1.fs, 
            ((rf1.fs * 100 / if(coalesce(rf2.fs, 0) = 0, 1, rf2.fs)) - 100) as growth 
        FROM (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM 
                request_frequency 
            WHERE 
                date >= toDate(%(v1)s) - 29 
            GROUP BY query_id
        ) as rf1 
        JOIN (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM request_frequency 
            WHERE 
                date BETWEEN toDate(%(v1)s) - 59 AND toDate(%(v1)s) - 30 
            GROUP BY query_id
        ) as rf2 
        ON rf1.query_id = rf2.query_id 
        ORDER BY growth DESC
    ) as rfn 
    ON rfn.query_id = r.id 
    JOIN (
        SELECT 
        rf3.query_id, 
        rf3.fs, 
        ((rf3.fs * 100 / if(coalesce(rf4.fs, 0) = 0, 1, rf4.fs)) - 100) as growth 
        FROM (
            SELECT 
            query_id, 
            sum(frequency) as fs 
            FROM 
                request_frequency 
            WHERE 
                date >= toDate(%(v1)s) - 59 
            GROUP BY 
                query_id
        ) as rf3 
        JOIN (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM 
                request_frequency 
            WHERE 
                date BETWEEN toDate(%(v1)s) - 119 AND toDate(%(v1)s) - 60 
            GROUP BY query_id
        ) as rf4 ON rf3.query_id = rf4.query_id 
        ORDER BY growth DESC
    ) as rfn60 ON rfn60.query_id = r.id 
    JOIN (
        SELECT 
            rf5.query_id, 
            rf5.fs, 
            ((rf5.fs * 100 / if(coalesce(rf6.fs, 0) = 0, 1, rf6.fs)) - 100) as growth
        FROM (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM 
                request_frequency 
            WHERE date >= toDate(%(v1)s) - 89 GROUP BY query_id
        ) as rf5 
        JOIN (
            SELECT 
                query_id, 
                sum(frequency) as fs 
            FROM 
                request_frequency 
            WHERE 
                date BETWEEN toDate(%(v1)s) - 179 AND toDate(%(v1)s) - 90 
            GROUP BY query_id
        ) as rf6 ON rf5.query_id = rf6.query_id 
        ORDER BY 
            growth DESC
    ) as rfn90 ON rfn90.query_id = r.id 
    WHERE
        rfn.growth >= 30
    AND
        rfn60.growth >= 30
    AND
        rfn90.growth >= 30
    ORDER BY 
        rfn.fs DESC LIMIT 500"""
    async with get_async_connection() as client:
        q = await client.query(stmt, parameters={"v1": date_})
        result = list(q.result_rows)
    dataset = []
    async with ClientSession() as http_session:
        queries = [row[0] for row in result]
        tasks = [
            asyncio.create_task(
                get_report_data(http_session=http_session, query_string=query)
            )
            for query in queries
        ]
        t_and_p = await asyncio.gather(*tasks)
        subjects_dict = await get_today_subjects_dict(http_session=http_session)
        for row, tp_row in zip(result, t_and_p):
            query = row[0]
            frequency = row[1]
            growth_30 = row[2]
            growth_60 = row[3]
            growth_90 = row[4]
            total, priority = tp_row
            dataset.append(
                (
                    query,
                    subjects_dict.get(priority, "Не определён"),
                    frequency,
                    total,
                    (frequency // total) if total else 0,
                    growth_30,
                    growth_60,
                    growth_90,
                )
            )
    return dataset


def create_file_from_dataset(date_, dataset: list[tuple]):
    today = date_
    wb = Workbook()
    ws = wb.active
    make_radar_header(
        ws=ws,
        sheet_title="Топ 500",
        name=f"Топ 500 запросов, которые росли в {MONTH_NAMES[today.month]} {today.year}г.",
    )

    ws["A4"].value = "Запрос"
    ws["B4"].value = "Приоритетный предмет"
    ws["C4"].value = "Частотность за 30 дней"
    ws["D4"].value = "Количество артикулов по запросу"
    ws["E4"].value = "Частотность на 1 артикул"
    ws["F4"].value = "Динамика за 30 дней, %"
    ws["G4"].value = "Динамика за 60 дней, %"
    ws["H4"].value = "Динамика за 90 дней, %"

    header_fill = PatternFill(
        start_color="a653ec", end_color="a653ec", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    column_widths = {
        "A": 35,  # Запрос
        "B": 25,  # Приоритетный предмет
        "C": 25,  # Частотность за 30 дней
        "D": 25,  # Количество артикулов
        "E": 25,  # Частотность на 1 артикул
        "F": 25,  # Динамика 30 дней
        "G": 25,  # Динамика 60 дней
        "H": 25,  # Динамика 90 дней
    }

    ws.row_dimensions[4].height = 40
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = f"A4:{get_column_letter(8)}4"  # Фильтры на заголовках
    for row in dataset:
        ws.append(row)

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    wb.close()
    return excel_stream


async def get_report_download_bytes(date_: date):
    dataset = await get_report_dataset(date_=date_)
    file = create_file_from_dataset(date_=date_, dataset=dataset)
    return file


def test():
    test_data = [(1, 1, 1, 1, 1, 1, 1, 1)]
    f = create_file_from_dataset(date_=date.today() - timedelta(days=2), dataset=test_data)
    with open("test.xlsx", "wb") as file:
        file.write(f.read())


if __name__ == "__main__":
    test()
